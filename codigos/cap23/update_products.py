# Python Journey - Chapter 23
# Corrects costs in produce sales spreadsheet.

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The produce types and their updated prices
PRICE_UPDATES = {'Garlic': 3.07,
                'Celery': 1.19,
                'Lemon': 1.27}

# Loop through the rows and update the prices.
for row in range(2, sheet.get_highest_row()): # skip the first row
    produce_name = sheet.cell(row = row, column = 1).value
    if produce_name in PRICE_UPDATES:
        sheet.cell(row = row, column = 2).value = PRICE_UPDATES[produce_name]

wb.save('update_produce_sales.xlsx')

