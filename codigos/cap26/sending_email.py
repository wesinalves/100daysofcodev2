# Python Journey - Chapter 26
# Sends emails based on payment status in spreadsheet.

import openpyxl, smtplib, sys

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

last_col = sheet.get_highest_column()
latest_month = sheet.cell(row = 1, column = last_col).value

# Check each member's payment status.
unpaid_menbers = {}
for r in range(2, sheet.get_highest_row() + 1):
    payment = sheet.cell(row = r, column = last_col).value
    if payment != 'paid':
        name = sheet.cell(row = r, column = 1).value
        email = sheet.cell(row = r, column = 2).value
        unpaid_menbers[name] = email
# Log in to email account.
smtp = smtplib.SMTP('smtp.gmail.com', 587)
smtp.ehlo()
smtp.starttls()
smtp.login('my_email', sys.argv[1])

# Send out reminder emails.
for name, email in unpaid_menbers.items():
    body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not \
paid dues for %s. Please make this payment as soon as possible. Thank you!'" % \
(latest_month, name, latest_month)

    print('Sending email to %s...' % email)
    status = smtp.sendmail('my_email', email, body)

    if status != {}:
        print('There was a problem sending email to %s: %s' % (email, status))
    
smtp.quit()